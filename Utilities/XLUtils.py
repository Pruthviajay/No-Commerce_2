import openpyxl
from openpyxl.styles import PatternFill

def Row_Count(file,Sheet_No):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[Sheet_No]
    return sheet.max_row

def Column_Count(file,Sheet_No):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[Sheet_No]
    return sheet.max_column

def Read_File(file,Sheet_No,Row_NO,Column_No):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[Sheet_No]
    return sheet.cell(Row_NO, Column_No).value

def write_File(file,Sheet_No,Row_No,Column_No,Data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[Sheet_No]
    sheet.cell(Row_No,Column_No).value=Data
    workbook.save(file)

def Green_Fill(file,Sheet_No,Row_No,Column_No):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[Sheet_No]
    Green_Color=PatternFill(start_color="60b212",end_color="60b212",
                            fill_type="solid")
    sheet.cell(Row_No,Column_No).fill=Green_Color

def Red_Fill(file,Sheet_No,Row_No,Column_No):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[Sheet_No]
    Red_Color=PatternFill(start_color="ff000",end_color="ff0000",
                            fill_type="solid")
    sheet.cell(Row_No,Column_No).fill=Red_Color