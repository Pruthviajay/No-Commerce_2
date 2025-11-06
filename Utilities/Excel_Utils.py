import openpyxl


# file="./Test_Data/Data.xlsx"
file=r"D:\pruthvi\Data science\Sdet_Pytest\Nopcommerce\Test_Data\Data.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]
row=sheet.max_row
column=sheet.max_column

def Row_Count(File,sheet):
    workbook=openpyxl.load_workbook(File)
    sheet=workbook[sheet]
    return sheet.max_row

def Column_Count(File,sheet):
    workbook=openpyxl.load_workbook(File)
    sheet=workbook[sheet]
    return sheet.max_row

def Read_Excel(File,sheet,Row,Column):
    Workbook=openpyxl.load_workbook(File)
    sheet=Workbook["Sheet1"]
    return sheet.cell(Row,Column).value

def Write_Excel(File,sheet,Row,Column,data):
    workbook=openpyxl.load_workbook(File)
    sheet=workbook[sheet]
    sheet.cell(Row,Column).value=data
    workbook.save(File)

# Write_Excel(file,"Sheet1",6,1,"adcd")
# print(Read_Excel(file,"Sheet1",6,1))


for r in range(2,(row+1)):
    for c in range(1,column+1):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook["Sheet1"]
        print(sheet.cell(r,c).value,end="    ")
    print()